import openpyxl
# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb.create_sheet('test_case')
# 保存为一个xlsx格式的文件
wb.save('cases.xlsx')
# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('cases.xlsx')
# 第二步：选取表单
sh = wb['Sheet1']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row = 1,column = 1)   # 读取第一行，第一列的数据
print(ce.value)
# 按行读取数据 list(sh.rows)
print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
for cases in list(sh.rows)[1:]:
    case_id =  cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted,case_data)
# 关闭工作薄
wb.close()


#####第二段

import openpyxl
class Case: #这个类用来存储用例的
    __slots__ = [] #特殊的类属性，可以用来限制这个类创建的实例属性添加 可写可不写
    pass

class ReadExcel(object): #读取excel数据的类
    def __init__(self,file_name,sheet_name):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sh = self.wb[sheet_name]
    def read_data_line(self):
        #按行读取数据转化为列表
        rows_data = list(self.sh.rows)
        # print(rows_data)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # print(titles)
        #定义一个空列表用来存储测试用例
        cases = []
        for case in rows_data[1:]:
            # print(case)
            data = []
            for cell in case: #获取一条测试用例数据
                # print(cell.value)
                data.append(cell.value)
                # print(data)
                #判断该单元格是否为字符串，如果是字符串类型则需要使用eval();如果不是字符串类型则不需要使用eval()
                if isinstance(cell.value,str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
                #将该条数据存放至cases中
            # print(dict(list(zip(titles,data))))
                case_data = dict(list(zip(titles,data)))
                cases.append(case_data)
        return cases
if __name__ == '__main__':
    r = ReadExcel('cases.xlsx','Sheet1')
    data1 = r.read_data_line()
    print(data1)
	
	
## 第三段

import openpyxl
class Case:
    pass
class ReadExcel(object):
    def __init__(self,filename,sheetname):
        self.wb = openpyxl.load_workbook(filename)
        self.sh = self.wb[sheetname]
    def read_data_obj(self):
        """
        按行读取数据  每条用例存储在一个对象中
        :return:
        """
        rows_data = list(self.sh.rows)
        # print(rows_data)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # print(titles)
        # 定义一个空列表用来存储测试用例
        cases = []
        for case in rows_data[1:]:
            # print(case)
            #创建一个Case类的对象，用来保存用例数据
            case_obj = Case()
            data = []
            for cell in case:  # 获取一条测试用例数据
                # print(cell.value)
                # data.append(cell.value)
                # print(data)
                if isinstance(cell.value,str):  # 判断该单元格是否为字符串，如果是字符串类型则需要使用eval();如果不是字符串类型则不需要使用eval()
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            # 将该条数据存放至cases中
            # print(dict(list(zip(titles,data))))
            case_data = list(zip(titles, data))
            # print(case_data)
            for i in case_data:
                setattr(case_obj,i[0],i[1])
            # print(case_obj)
            # print(case_obj.case_id,case_obj.data,case_obj.excepted)
            cases.append(case_obj)
        return cases
if  __name__ == '__main__':
    r = ReadExcel('cases.xlsx','Sheet1')
    res = r.read_data_obj()
    for i in res:
        print(i.caseid, i.excepted, i.data)
		
### 第四段

import openpyxl
class Case:
    pass
class ReadExcelZy(object):
    def __init__(self,filename,sheetname):
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb[sheetname]
        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在字典中
        # 所有的用例放在列表中并且进行返回
    def read_data(self,list1):
        """
        :param list1:  list--->要读取列   list类型
        :return:    返回一个列表，每一个元素为一个用例（用例为dict类型）
        """
        # 获取最大的行数
        max_r = self.sheet.max_row
        cases = []   #定义一个空列表，用来存放所有的用例数据
        titles = []   #定义一个空列表，用来存放表头
        # 遍历所有的行数据
        for row in range(1,max_r+1):
            if row != 1:      #判断是否是第一行
                case_data = [] #定义一个空列表，用来存放该行的用例数据
                for column in list1:
                    info = self.sheet.cell(row,column).value
                    # print(info)
                    case_data.append(info)
                    # print(list(zip(titles,case_data)))
                case = dict(zip(titles,case_data))  #将该条数据和表头进行打包组合,作用相当于dict(list(zip(titles,case_data)))
                # print(case)
                cases.append(case)
                # print(cases)
            else:   #获取表头数据
                for column in list1:
                    title = self.sheet.cell(row,column).value
                    titles.append(title)
                # print(titles)
        return cases
if __name__ == '__main__':
    r = ReadExcelZy("cases.xlsx","Sheet1")
    res = r.read_data([1,2,3])
    for o in res:
        print(o['caseid'],o['data'],o['excepted'])
		
##第五段

import openpyxl
class Case:
    pass
class ReadExcelZy(object):
    def __init__(self,filename,sheetname):
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb[sheetname]

        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在字典中
        # 所有的用例放在对象中并且进行返回
    def read_data_obj(self,list2):
        max_r1 = self.sheet.max_row      #获取最大行数
        cases = []
        titles = []      #用来存放表头数据
        for row in range(1,max_r1+1):
            if row != 1:
                case_data = []
                for column in list2:
                    info = self.sheet.cell(row,column).value
                    # print(info)
                    case_data.append(info)
                cases_data = list(zip(titles,case_data))
                #将一条用例存到一个对象中（每一列对应对象的一个属性）
                case_obj = Case()
                for i in cases_data:
                    # print(i)
                    setattr(case_obj,i[0],i[1])
                # print(case_obj.caseid,case_obj.excepted,case_obj.data)
                cases.append(case_obj)
            else:
                for column in list2:
                    title = self.sheet.cell(row,column).value
                    titles.append(title)
        return cases
if __name__ == '__main__':
    r = ReadExcelZy("cases.xlsx","Sheet1")
    res = r.read_data_obj([1,2,3])
    for i in res:
        print(i.caseid,i.data,i.excepted)
		
## 第六段

import openpyxl
class Case:  # 这个类用来存储用例的
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs:zip类型——>[{key,value},(key1,value1)......]
        """
        for i in attrs:
            setattr(self, i[0], i[1])
class ReadExcel(object):
    def __init__(self, filename, sheetname):
        """
        定义需要打开的文件及表名
        :param filename:   文件名
        :param sheetname:  表名
        """
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb[sheetname]
    def read_data_obj_new(self, list2):
        # 获取最大行数
        max_r1 = self.sheet.max_row
        cases = []
        # 用来存放表头数据
        titles = []
        for row in range(1, max_r1 + 1):
            if row != 1:
                case_data = []
                for column in list2:
                    info = self.sheet.cell(row, column).value
                    # print(info)
                    case_data.append(info)
                case = list(zip(titles, case_data))
                # 新建对象时，将对象传给Case类
                case_obj = Case(case)
                # print(case_obj.caseid,case_obj.excepted,case_obj.data)
                cases.append(case_obj)
            else:
                # 获取表头
                for column in list2:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
                if None in titles:
                    raise ValueError("传入的表头的数据有显示为空")
        return cases
if __name__ == '__main__':
    r = ReadExcel('cases.xlsx', 'Sheet1')
    res1 = r.read_data_obj_new([1, 2, 3])
    for i in res1:
        print(i.caseid, i.data, i.excepted)
		
## 第七段 

import unittest
from python.register_new.register import register
from python.register_new.register_testcase_new import RegisterTestCase
from HTMLTestRunnerNew import HTMLTestRunner
class RegisterTestCase(unittest.TestCase):
    # 初始化测试用例
    def __init__(self,modethod_name,excepted,data):
        # modethod_name 测试用例方法名
        super().__init__(modethod_name)
        # excepted 测试用例的预期结果
        self.excepted = excepted
        # data 测试用例参数值
        self.data = data

    def setUp(self):
        print("准备测试环境，执行测试用例之前会执行此操作")

    def tearDown(self):
        print("还原测试环境，执行完测试用例之后会执行此操作")

    def test_register(self):
        res = register(*self.data)
        try:
            self.assertEquals(self.excepted,res)
        except AssertionError as e:
            print("该条测试用例执行未通通过")
            raise e
        else:
            print("该条测试用例执行通过")

# 创建测试套件
suite = unittest.TestSuite()

# 将测试用例添加至测试套件中
case = [{'excepted':'{"code": 1, "msg": "注册成功"}','data':'('python1', '123456','123456')'},
        {'excepted':'{"code": 0, "msg": "两次密码不一致"}','data':'('python1', '1234567','123456')'}]
for case in cases:
    suite.addTest(RegisterTestCase('test_register',case['excepted'],case['data']))

# 执行测试套件，生成测试报告
with open("report.html",'wb') as f:
    runner = HTMLTestRunner(
        stream = f,
        verbosity = 2,
        title = 'python_test_report',
        description = '这是一份测试报告',
        tester = 'WL'
    )
    runner.run(suite)

## 第八段

import unittest
from python.register_new.register import register
from python.register_new.register_testcase_new import RegisterTestCase
from HTMLTestRunnerNew import HTMLTestRunner
from python.readexcel import ReadExcel


class RegisterTestCase(unittest.TestCase):
    # 初始化测试用例
    def __init__(self, modethod_name, excepted, data):
        # modethod_name 测试用例方法名
        super().__init__(modethod_name)
        # excepted 测试用例的预期结果
        self.excepted = excepted
        # data 测试用例参数值
        self.data = data

    def setUp(self):
        print("准备测试环境，执行测试用例之前会执行此操作")

    def tearDown(self):
        print("还原测试环境，执行完测试用例之后会执行此操作")

    def test_register(self):
        res = register(*self.data)
        try:
            self.assertEquals(self.excepted, res)
        except AssertionError as e:
            print("该条测试用例执行未通通过")
            raise e
        else:
            print("该条测试用例执行通过")


# 创建测试套件
suite = unittest.TestSuite()
# 调用封装好的读取数据的Excel类，获取测试数据
r1 = ReadExcel('cases.xlsx', 'Sheet1')
cases = r1.read_data_obj_new([2, 3])
# 将测试用例添加至测试套件中
for case in cases:
    # 需要使用eval()函数对except和data进行自动识别
    suite.addTest(RegisterTestCase('test_register', eval(case.excepted), eval(case.data)))
# 执行测试套件，生成测试报告
with open("report.html", 'wb') as f:
    runner = HTMLTestRunner(
        stream=f,
        verbosity=2,
        title='python_test_report',
        description='这是一份测试报告',
        tester='WL')
    runner.run(suite)
	
## end

